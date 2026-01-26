#!/usr/bin/env python3
"""
测试Excel导入功能 - 不连接数据库，只解析Excel
"""

import sys
from pathlib import Path
from datetime import datetime, timedelta

try:
    from openpyxl import load_workbook
except ImportError:
    print("请安装openpyxl: pip install openpyxl")
    sys.exit(1)


def test_parse_excel(excel_path: str):
    """测试解析Excel文件"""
    
    print(f"开始解析Excel文件: {excel_path}")
    
    # 读取Excel文件
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        print(f"成功打开Excel文件，工作表: {ws.title}")
    except Exception as e:
        print(f"读取Excel文件失败: {e}")
        return False
    
    # 读取所有行
    rows = list(ws.iter_rows(values_only=True))
    
    if len(rows) < 3:
        print("Excel文件数据不足")
        return False
    
    # 解析第3行：客户信息行
    client_info_row = rows[2]
    
    client_name = str(client_info_row[1]).strip() if client_info_row[1] else None
    case_type = str(client_info_row[2]).strip() if client_info_row[2] else "GTV"
    start_date_str = str(client_info_row[4]).strip() if client_info_row[4] else None
    
    print(f"\n客户信息:")
    print(f"  姓名: {client_name}")
    print(f"  类型: {case_type}")
    print(f"  启动时间: {start_date_str}")
    
    # 解析启动时间
    start_date_parsed = None
    if start_date_str:
        try:
            start_date_parsed = datetime.strptime(start_date_str, "%Y-%m-%d %H:%M:%S")
        except:
            try:
                start_date_parsed = datetime.strptime(start_date_str, "%Y-%m-%d")
            except:
                pass
    
    print(f"\n时间规划:")
    
    # 解析时间规划（从第5行开始）
    current_task = None
    current_description = None
    task_count = 0
    
    for row_idx, row in enumerate(rows[4:], start=5):
        if not any(row):
            continue
        
        row_values = [str(cell).strip() if cell else "" for cell in row]
        
        # 检查是否是时间段行（包含"第X个月"）
        time_period = row_values[0] if row_values[0] else ""
        task_name = row_values[1] if len(row_values) > 1 and row_values[1] else ""
        note = row_values[4] if len(row_values) > 4 and row_values[4] else ""
        
        # 如果是时间段行
        if "第" in time_period and "月" in time_period:
            current_task = task_name if task_name else time_period
            current_description = ""
            print(f"\n  [{time_period}] {current_task}")
            if note:
                print(f"    备注: {note}")
        
        # 如果是描述行（第一列为空，第二列有内容）
        elif not row_values[0] and row_values[1]:
            current_description = row_values[1]
            print(f"    描述: {current_description[:100]}...")
            if note:
                print(f"    备注: {note}")
        
        # 如果是预计完成时间行
        elif "预计完成时间" in row_values[0] or "完成时间" in row_values[0]:
            if current_task and start_date_str:
                # 解析公式计算日期
                formula = row_values[1] if len(row_values) > 1 else ""
                due_date = None
                
                if formula.startswith("="):
                    # 解析公式中的天数
                    if "+" in formula:
                        try:
                            days_str = formula.split("+")[-1].strip()
                            days = int(days_str)
                            start_dt = datetime.strptime(start_date_str, "%Y-%m-%d %H:%M:%S")
                            due_date = (start_dt + timedelta(days=days)).date()
                            print(f"    预计完成时间: {due_date.isoformat()} (启动时间 + {days}天)")
                            task_count += 1
                        except Exception as e:
                            print(f"    解析日期公式失败: {formula}, {e}")
                
                # 重置当前任务
                current_task = None
                current_description = None
    
    print(f"\n总共解析到 {task_count} 个任务")
    print("\n解析完成！如果配置了Supabase环境变量，可以运行 import_excel_data.py 导入数据库")
    
    return True


if __name__ == "__main__":
    excel_path = "../副本英国GTV文案进度表-李成(1).xlsx"
    
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    
    if not Path(excel_path).exists():
        print(f"文件不存在: {excel_path}")
        sys.exit(1)
    
    success = test_parse_excel(excel_path)
    sys.exit(0 if success else 1)

