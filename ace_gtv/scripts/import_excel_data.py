#!/usr/bin/env python3
"""
从Excel文件导入客户和案件数据到数据库
"""

import os
import sys
from pathlib import Path
from datetime import datetime, timedelta
import json

# 添加当前目录到路径
sys.path.append(os.path.dirname(__file__))

try:
    from openpyxl import load_workbook
except ImportError:
    print("请安装openpyxl: pip install openpyxl")
    sys.exit(1)

from document_management_db import DocumentManagementDB
from logger_config import setup_module_logger

logger = setup_module_logger("import_excel_data", os.getenv("LOG_LEVEL", "INFO"))


def parse_date(date_str):
    """解析日期字符串"""
    if not date_str:
        return None
    
    # 尝试多种日期格式
    date_formats = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%m/%d/%Y",
        "%d/%m/%Y",
    ]
    
    for fmt in date_formats:
        try:
            return datetime.strptime(str(date_str).strip(), fmt).date().isoformat()
        except:
            continue
    
    # 如果是datetime对象
    if isinstance(date_str, datetime):
        return date_str.date().isoformat()
    
    return str(date_str)


def calculate_date_from_formula(formula: str, start_date_str: str):
    """从Excel公式计算日期"""
    try:
        # 解析启动时间
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d %H:%M:%S")
        
        # 解析公式，如 "=E3+30" 表示从E3单元格的日期加30天
        if "+" in formula:
            parts = formula.split("+")
            days = int(parts[-1].strip())
            return (start_date + timedelta(days=days)).date().isoformat()
        elif "-" in formula:
            parts = formula.split("-")
            days = int(parts[-1].strip())
            return (start_date - timedelta(days=days)).date().isoformat()
    except Exception as e:
        logger.warning(f"解析日期公式失败: {formula}, {e}")
    return None


def import_excel_to_database(excel_path: str):
    """从Excel文件导入数据到数据库 - 针对GTV文案进度表格式"""
    
    logger.info(f"开始导入Excel文件: {excel_path}")
    
    # 初始化数据库
    try:
        db = DocumentManagementDB()
        logger.info("数据库连接成功")
    except Exception as e:
        logger.error(f"数据库连接失败: {e}")
        return False
    
    # 读取Excel文件
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        logger.info(f"成功打开Excel文件，工作表: {ws.title}")
    except Exception as e:
        logger.error(f"读取Excel文件失败: {e}")
        return False
    
    # 读取所有行
    rows = list(ws.iter_rows(values_only=True))
    
    if len(rows) < 3:
        logger.warning("Excel文件数据不足")
        return False
    
    # 解析第3行：客户信息行
    # 格式: ['客户姓名', '李成', '数字科技类', '启动时间', '2025-11-05 00:00:00', '']
    client_info_row = rows[2]  # 第3行，索引为2
    
    client_name = str(client_info_row[1]).strip() if client_info_row[1] else None
    case_type = str(client_info_row[2]).strip() if client_info_row[2] else "GTV"
    start_date_str = str(client_info_row[4]).strip() if client_info_row[4] else None
    
    if not client_name:
        logger.error("未找到客户姓名")
        return False
    
    logger.info(f"客户信息: 姓名={client_name}, 类型={case_type}, 启动时间={start_date_str}")
    
    # 检查或创建客户
    clients_result = db.get_clients(limit=1000)
    existing_client = None
    if clients_result.get("success"):
        for client in clients_result.get("data", []):
            if client.get("name") == client_name:
                existing_client = client
                break
    
    if existing_client:
        client_id = existing_client["id"]
        logger.info(f"找到已存在客户: {client_name} (ID: {client_id})")
    else:
        client_data = {
            "name": client_name,
            "email": "",
            "phone": "",
            "nationality": "",
            "passport_number": "",
            "notes": f"从Excel导入 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        }
        result = db.create_client(client_data)
        if result["success"]:
            client_id = result["data"]["id"]
            logger.info(f"创建新客户: {client_name} (ID: {client_id})")
        else:
            logger.error(f"创建客户失败: {result.get('error')}")
            return False
    
    # 创建案件
    start_date_parsed = None
    if start_date_str:
        try:
            start_date_parsed = datetime.strptime(start_date_str, "%Y-%m-%d %H:%M:%S").date().isoformat()
        except:
            try:
                start_date_parsed = datetime.strptime(start_date_str, "%Y-%m-%d").date().isoformat()
            except:
                pass
    
    case_data = {
        "client_id": client_id,
        "case_type": case_type if case_type else "GTV",
        "visa_type": "GTV",
        "status": "draft",  # 使用draft作为默认状态（enum可能不支持in_progress）
        "priority": "normal",
        "description": f"GTV签证申请 - {case_type}",
        "target_submission_date": None,  # 可以根据最后一个任务计算
    }
    
    result = db.create_case(case_data)
    if not result["success"]:
        logger.error(f"创建案件失败: {result.get('error')}")
        return False
    
    case_id = result["data"]["id"]
    logger.info(f"创建案件: {case_type} (ID: {case_id})")
    
    # 解析时间规划（从第5行开始）
    current_task = None
    current_description = None
    current_note = None
    
    for row_idx, row in enumerate(rows[4:], start=5):  # 从第5行开始
        if not any(row):
            continue
        
        row_values = [str(cell).strip() if cell else "" for cell in row]
        
        # 检查是否是时间段行（包含"第X个月"）
        time_period = row_values[0] if row_values[0] else ""
        task_name = row_values[1] if len(row_values) > 1 and row_values[1] else ""
        note = row_values[4] if len(row_values) > 4 and row_values[4] else ""
        
        # 如果是时间段行
        if "第" in time_period and "月" in time_period:
            current_task = task_name if task_name else time_period
            current_description = ""
            current_note = note
        
        # 如果是描述行（第一列为空，第二列有内容）
        elif not row_values[0] and row_values[1]:
            current_description = row_values[1]
            if note:
                current_note = note
        
        # 如果是预计完成时间行
        elif "预计完成时间" in row_values[0] or "完成时间" in row_values[0]:
            if current_task and start_date_str:
                # 解析公式计算日期
                formula = row_values[1] if len(row_values) > 1 else ""
                due_date = None
                
                if formula.startswith("="):
                    # 解析公式中的天数
                    if "+" in formula:
                        try:
                            days_str = formula.split("+")[-1].strip()
                            days = int(days_str)
                            start_dt = datetime.strptime(start_date_str, "%Y-%m-%d %H:%M:%S")
                            due_date = (start_dt + timedelta(days=days)).date().isoformat()
                        except:
                            pass
                
                # 创建时间规划
                if current_task:
                    timeline_data = {
                        "case_id": case_id,
                        "task_name": current_task,
                        "task_type": "document",
                        "start_date": start_date_parsed,
                        "due_date": due_date,
                        "status": "pending",
                        "description": current_description if current_description else current_task,
                        "assignee": "",
                    }
                    
                    result = db.create_timeline(timeline_data)
                    if result["success"]:
                        logger.info(f"创建时间规划: {current_task} (截止: {due_date})")
                    elif result.get("skip"):
                        logger.info(f"跳过时间规划（表结构不匹配）: {current_task}")
                    else:
                        logger.warning(f"创建时间规划失败: {result.get('error')}")
                    
                    # 同时创建进度记录
                    progress_data = {
                        "case_id": case_id,
                        "milestone": current_task,
                        "status": "pending",
                        "description": current_description if current_description else current_task,
                    }
                    
                    result = db.create_progress(progress_data)
                    if result["success"]:
                        logger.info(f"创建进度记录: {current_task}")
                    
                    # 重置当前任务
                    current_task = None
                    current_description = None
    
    logger.info("Excel数据导入完成")
    return True


if __name__ == "__main__":
    excel_path = "../副本英国GTV文案进度表-李成(1).xlsx"
    
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    
    if not os.path.exists(excel_path):
        logger.error(f"文件不存在: {excel_path}")
        sys.exit(1)
    
    success = import_excel_to_database(excel_path)
    sys.exit(0 if success else 1)

